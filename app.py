from flask import Flask, render_template, request, jsonify, g, send_file
from dotenv import load_dotenv
from werkzeug.utils import secure_filename
from werkzeug.middleware.proxy_fix import ProxyFix
import pandas as pd
import os
import sqlite3
import json
import hmac
import io
import csv
import hashlib
import threading
import zipfile
import uuid
import time
from datetime import datetime, timezone, timedelta
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor
import joblib
import logging
import platform
import sklearn
from atlas import ai
from atlas.config import Settings
from atlas.security import InMemoryRateLimiter, add_security_headers
from sklearn.compose import ColumnTransformer
from sklearn.dummy import DummyClassifier, DummyRegressor
from sklearn.ensemble import RandomForestClassifier, RandomForestRegressor
from sklearn.impute import SimpleImputer
from sklearn.linear_model import LinearRegression, LogisticRegression
from sklearn.metrics import (accuracy_score, f1_score, precision_score, recall_score,
                             mean_absolute_error, mean_squared_error, r2_score,
                             confusion_matrix)
from sklearn.model_selection import train_test_split, cross_val_score, StratifiedKFold, KFold
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import OneHotEncoder, StandardScaler
from openpyxl import load_workbook
import re

load_dotenv()
settings = Settings.from_env()
logging.basicConfig(
    level=os.getenv("LOG_LEVEL", "INFO"),
    format="%(asctime)s %(levelname)s %(name)s %(message)s",
)
logger = logging.getLogger("atlas")

app = Flask(__name__)
if os.getenv("RENDER"):
    app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)
app.config["JSON_AS_ASCII"] = False
os.makedirs(app.instance_path, exist_ok=True)
DATA_FOLDER = os.path.abspath(os.getenv("ATLAS_DATA_DIR") or app.instance_path)
BANCO_DADOS = os.path.join(DATA_FOLDER, "conversas.db")

UPLOAD_FOLDER = os.path.join(DATA_FOLDER, "uploads")
MODEL_FOLDER = os.path.join(DATA_FOLDER, "modelos")
PREDICTION_FOLDER = os.path.join(DATA_FOLDER, "previsoes")
EXTENSOES_PERMITIDAS = {"csv", "xlsx"}
TAMANHO_MAXIMO = settings.max_upload_bytes

app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config["MAX_CONTENT_LENGTH"] = TAMANHO_MAXIMO

os.makedirs(DATA_FOLDER, exist_ok=True)
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(MODEL_FOLDER, exist_ok=True)
os.makedirs(PREDICTION_FOLDER, exist_ok=True)
executor = ThreadPoolExecutor(max_workers=2, thread_name_prefix="atlas-ml")
rate_limiter = InMemoryRateLimiter(settings.rate_limit_per_minute)
task_slots = threading.BoundedSemaphore(settings.max_pending_tasks)


@app.after_request
def cabecalhos_seguranca(resposta):
    return add_security_headers(resposta)


@app.before_request
def limitar_requisicoes():
    if request.endpoint in {"inicio", "static", "saude"}:
        return None
    return rate_limiter.check()

def obter_banco():
    if "banco" not in g:
        g.banco = sqlite3.connect(BANCO_DADOS)
        g.banco.row_factory = sqlite3.Row
        g.banco.execute("PRAGMA foreign_keys = ON")
        g.banco.execute("PRAGMA journal_mode = WAL")
        g.banco.execute("PRAGMA busy_timeout = 5000")
    return g.banco


@app.teardown_appcontext
def fechar_banco(erro=None):
    banco = g.pop("banco", None)
    if banco is not None:
        banco.close()


def iniciar_banco():
    with app.app_context():
        banco = obter_banco()
        banco.executescript("""
            CREATE TABLE IF NOT EXISTS schema_migrations (
                versao INTEGER PRIMARY KEY,
                aplicada_em TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS conversas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                titulo TEXT NOT NULL,
                criada_em TEXT NOT NULL,
                atualizada_em TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS mensagens (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                conversa_id INTEGER NOT NULL,
                autor TEXT NOT NULL CHECK (autor IN ('usuario', 'agente')),
                conteudo TEXT NOT NULL,
                criada_em TEXT NOT NULL,
                FOREIGN KEY (conversa_id) REFERENCES conversas(id) ON DELETE CASCADE
            );
            CREATE INDEX IF NOT EXISTS idx_mensagens_conversa
                ON mensagens(conversa_id, id);
            CREATE TABLE IF NOT EXISTS arquivos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                conversa_id INTEGER NOT NULL,
                nome_original TEXT NOT NULL,
                caminho TEXT NOT NULL UNIQUE,
                extensao TEXT NOT NULL,
                tamanho INTEGER NOT NULL,
                linhas INTEGER NOT NULL,
                colunas INTEGER NOT NULL,
                resumo_json TEXT NOT NULL,
                criada_em TEXT NOT NULL,
                FOREIGN KEY (conversa_id) REFERENCES conversas(id) ON DELETE CASCADE
            );
            CREATE TABLE IF NOT EXISTS modelos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                conversa_id INTEGER NOT NULL,
                arquivo_id INTEGER NOT NULL,
                alvo TEXT NOT NULL,
                tipo TEXT NOT NULL,
                nome TEXT NOT NULL,
                caminho TEXT NOT NULL UNIQUE,
                recursos_json TEXT NOT NULL,
                metricas_json TEXT NOT NULL,
                criada_em TEXT NOT NULL,
                FOREIGN KEY (conversa_id) REFERENCES conversas(id) ON DELETE CASCADE,
                FOREIGN KEY (arquivo_id) REFERENCES arquivos(id) ON DELETE CASCADE
            );
            CREATE TABLE IF NOT EXISTS tarefas (
                id TEXT PRIMARY KEY,
                conversa_id INTEGER NOT NULL,
                tipo TEXT NOT NULL,
                status TEXT NOT NULL,
                progresso INTEGER NOT NULL DEFAULT 0,
                resultado_json TEXT,
                erro TEXT,
                criada_em TEXT NOT NULL,
                atualizada_em TEXT NOT NULL,
                FOREIGN KEY (conversa_id) REFERENCES conversas(id) ON DELETE CASCADE
            );
        """)
        colunas_mensagens = {item[1] for item in banco.execute("PRAGMA table_info(mensagens)")}
        if "status" not in colunas_mensagens:
            banco.execute("ALTER TABLE mensagens ADD COLUMN status TEXT NOT NULL DEFAULT 'concluida'")
        if "erro" not in colunas_mensagens:
            banco.execute("ALTER TABLE mensagens ADD COLUMN erro TEXT")
        colunas_tarefas = {item[1] for item in banco.execute("PRAGMA table_info(tarefas)")}
        if "requisicao_json" not in colunas_tarefas:
            banco.execute("ALTER TABLE tarefas ADD COLUMN requisicao_json TEXT")
        colunas_modelos = {item[1] for item in banco.execute("PRAGMA table_info(modelos)")}
        if "sha256" not in colunas_modelos:
            banco.execute("ALTER TABLE modelos ADD COLUMN sha256 TEXT")
        banco.execute("INSERT OR IGNORE INTO schema_migrations(versao,aplicada_em) VALUES(2,?)", (agora_iso(),))
        banco.commit()


def agora_iso():
    return datetime.now(timezone.utc).isoformat()


def salvar_mensagem(conversa_id, autor, conteudo, status="concluida", erro=None):
    banco = obter_banco()
    if banco.execute("SELECT id FROM conversas WHERE id = ?", (conversa_id,)).fetchone() is None:
        raise ValueError("Conversa não encontrada.")
    agora = agora_iso()
    cursor = banco.execute(
        "INSERT INTO mensagens (conversa_id, autor, conteudo, criada_em, status, erro) VALUES (?, ?, ?, ?, ?, ?)",
        (conversa_id, autor, conteudo, agora, status, erro)
    )
    banco.execute("UPDATE conversas SET atualizada_em = ? WHERE id = ?", (agora, conversa_id))
    banco.commit()
    return cursor.lastrowid


def obter_conversa(conversa_id, banco=None):
    banco = banco or obter_banco()
    return banco.execute("SELECT * FROM conversas WHERE id=?", (conversa_id,)).fetchone()


def arquivo_da_conversa(conversa_id, banco=None):
    banco = banco or obter_banco()
    conversa = obter_conversa(conversa_id, banco)
    if conversa is None:
        return None
    return banco.execute(
        "SELECT * FROM arquivos WHERE conversa_id = ? ORDER BY id DESC LIMIT 1", (conversa_id,)
    ).fetchone()


def modelo_da_conversa(conversa_id, banco=None):
    banco = banco or obter_banco()
    conversa = obter_conversa(conversa_id, banco)
    if conversa is None:
        return None
    return banco.execute(
        "SELECT * FROM modelos WHERE conversa_id = ? ORDER BY id DESC LIMIT 1", (conversa_id,)
    ).fetchone()


def carregar_dataframe(registro):
    caminho = registro["caminho"]
    if not os.path.isfile(caminho):
        raise ValueError("O arquivo associado não está mais disponível.")
    return ler_dataframe(caminho, registro["extensao"])


def ler_dataframe(fonte, extensao):
    if extensao != "csv":
        caminho = fonte if isinstance(fonte, (str, os.PathLike)) else None
        if caminho:
            try:
                with zipfile.ZipFile(caminho) as pacote:
                    total = sum(item.file_size for item in pacote.infolist())
                    comprimido = max(1, sum(item.compress_size for item in pacote.infolist()))
                    if total > settings.max_upload_bytes * 20 or total / comprimido > 100:
                        raise ValueError("A planilha compactada expande além do limite seguro.")
            except zipfile.BadZipFile as erro:
                raise ValueError("O arquivo XLSX está corrompido.") from erro
            livro = load_workbook(caminho, read_only=True, data_only=True)
            try:
                for planilha in livro.worksheets:
                    if planilha.max_row > settings.max_rows + 1 or planilha.max_column > settings.max_columns:
                        raise ValueError("A planilha excede os limites de linhas ou colunas.")
            finally:
                livro.close()
        return pd.read_excel(fonte)
    if hasattr(fonte, "read"):
        bruto = fonte.read()
    else:
        bruto = Path(fonte).read_bytes()
    if not bruto.strip():
        raise pd.errors.EmptyDataError("O arquivo está vazio.")
    ultimo_erro = None
    for codificacao in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return pd.read_csv(
                io.BytesIO(bruto), encoding=codificacao, sep=None, engine="python"
            )
        except (UnicodeDecodeError, pd.errors.ParserError, csv.Error) as erro:
            ultimo_erro = erro
    raise ValueError("Não foi possível identificar a codificação ou separador do CSV.") from ultimo_erro


def remover_arquivo_seguro(caminho, pasta_permitida):
    alvo = Path(caminho).resolve()
    raiz = Path(pasta_permitida).resolve()
    if raiz not in alvo.parents or not alvo.is_file() or alvo.is_symlink():
        return
    alvo.unlink()


def limpar_previsoes_expiradas():
    limite = datetime.now(timezone.utc) - timedelta(hours=settings.artifact_ttl_hours)
    raiz = Path(PREDICTION_FOLDER).resolve()
    for item in raiz.glob("*.csv"):
        if item.is_symlink() or not item.is_file():
            continue
        alterado = datetime.fromtimestamp(item.stat().st_mtime, timezone.utc)
        if alterado < limite:
            remover_arquivo_seguro(str(item), PREDICTION_FOLDER)
def sha256_arquivo(caminho):
    digest = hashlib.sha256()
    with open(caminho, "rb") as arquivo:
        for bloco in iter(lambda: arquivo.read(1024 * 1024), b""):
            digest.update(bloco)
    return digest.hexdigest()


def neutralizar_formulas_csv(df):
    saida = df.copy()
    for coluna in saida.select_dtypes(include=["object", "string"]).columns:
        saida[coluna] = saida[coluna].map(
            lambda valor: "'" + valor if isinstance(valor, str) and valor.lstrip().startswith(("=", "+", "-", "@")) else valor
        )
    return saida


PADRAO_SENSIVEL = re.compile(r"(cpf|cnpj|e.?mail|email|telefone|celular|senha|token|cart[aã]o|conta|endere[cç]o|nome)", re.I)


def dataframe_privado_para_prompt(df):
    renomear = {}
    for indice, coluna in enumerate(df.columns, 1):
        if PADRAO_SENSIVEL.search(str(coluna)):
            renomear[coluna] = f"[coluna_sensivel_{indice}]"
    return df.rename(columns=renomear), [str(c) for c in renomear]


@app.route("/conversas", methods=["GET"])
def listar_conversas():
    linhas = obter_banco().execute("""
        SELECT c.id, c.titulo, c.criada_em, c.atualizada_em, COUNT(m.id) AS total_mensagens
        FROM conversas c LEFT JOIN mensagens m ON m.conversa_id = c.id
        GROUP BY c.id ORDER BY c.atualizada_em DESC
    """).fetchall()
    return jsonify([dict(linha) for linha in linhas])


@app.route("/conversas", methods=["POST"])
def criar_conversa():
    agora = agora_iso()
    banco = obter_banco()
    cursor = banco.execute(
        "INSERT INTO conversas (titulo, criada_em, atualizada_em) VALUES (?, ?, ?)",
        ("Nova conversa", agora, agora)
    )
    banco.commit()
    return jsonify({"id": cursor.lastrowid, "titulo": "Nova conversa"}), 201


@app.route("/conversas/<int:conversa_id>/mensagens", methods=["GET"])
def listar_mensagens(conversa_id):
    banco = obter_banco()
    conversa = banco.execute("SELECT id, titulo FROM conversas WHERE id=?", (conversa_id,)).fetchone()
    if conversa is None:
        return jsonify({"erro": "Conversa não encontrada."}), 404
    mensagens = banco.execute(
        "SELECT id, autor, conteudo, criada_em, status, erro FROM mensagens WHERE conversa_id = ? ORDER BY id",
        (conversa_id,)
    ).fetchall()
    arquivo = arquivo_da_conversa(conversa_id, banco)
    modelo = modelo_da_conversa(conversa_id, banco)
    arquivo_publico = None
    if arquivo:
        arquivo_publico = {
            "id": arquivo["id"], "nome_original": arquivo["nome_original"],
            "tamanho": arquivo["tamanho"], "linhas": arquivo["linhas"],
            "colunas": arquivo["colunas"], "resumo_json": arquivo["resumo_json"],
            "criada_em": arquivo["criada_em"]
        }
    modelo_publico = None
    if modelo:
        modelo_publico = {
            "id": modelo["id"], "alvo": modelo["alvo"], "tipo": modelo["tipo"],
            "nome": modelo["nome"], "metricas_json": modelo["metricas_json"],
            "criada_em": modelo["criada_em"]
        }
    return jsonify({
        "conversa": dict(conversa),
        "mensagens": [dict(m) for m in mensagens],
        "arquivo": arquivo_publico,
        "modelo": modelo_publico
    })


@app.route("/conversas/<int:conversa_id>", methods=["DELETE"])
def excluir_conversa(conversa_id):
    banco = obter_banco()
    ativa = banco.execute("SELECT 1 FROM tarefas WHERE conversa_id=? AND status IN ('pendente','processando')", (conversa_id,)).fetchone()
    if ativa: return jsonify({"erro": "Cancele ou aguarde o treinamento antes de excluir a conversa."}), 409
    arquivos = banco.execute("SELECT caminho FROM arquivos WHERE conversa_id=?", (conversa_id,)).fetchall()
    modelos = banco.execute("SELECT caminho FROM modelos WHERE conversa_id=?", (conversa_id,)).fetchall()
    cursor = banco.execute("DELETE FROM conversas WHERE id=?", (conversa_id,))
    banco.commit()
    if cursor.rowcount == 0:
        return jsonify({"erro": "Conversa não encontrada."}), 404
    for item in arquivos:
        remover_arquivo_seguro(item["caminho"], UPLOAD_FOLDER)
    for item in modelos:
        remover_arquivo_seguro(item["caminho"], MODEL_FOLDER)
    return "", 204


def extensao_permitida(nome):
    return (
        "." in nome
        and nome.rsplit(".", 1)[1].lower() in EXTENSOES_PERMITIDAS
    )


def gerar_resumo_dataframe(df):
    resumo = []

    resumo.append(f"Total de linhas: {len(df)}")
    resumo.append(f"Total de colunas: {len(df.columns)}")

    resumo.append(
        "Colunas: " + ", ".join(map(str, df.columns))
    )

    resumo.append("\nTipos das colunas:")

    for coluna, tipo in df.dtypes.items():
        resumo.append(
            f"- {coluna}: {tipo}"
        )

    resumo.append("\nValores nulos:")

    nulos = df.isnull().sum()

    for coluna, quantidade in nulos.items():
        if quantidade > 0:
            resumo.append(
                f"- {coluna}: {int(quantidade)}"
            )

    resumo.append(
        f"\nLinhas duplicadas: {int(df.duplicated().sum())}"
    )

    # Estatísticas das colunas numéricas
    numericas = df.select_dtypes(
        include="number"
    )

    if not numericas.empty:

        resumo.append(
            "\nEstatísticas das colunas numéricas:"
        )

        estatisticas = numericas.describe()

        resumo.append(
            estatisticas.to_string()
        )

    # Principais valores em colunas categóricas
    categoricas = df.select_dtypes(
        include=["object", "category"]
    )

    if not categoricas.empty and settings.send_data_samples:

        resumo.append(
            "\nPrincipais valores das colunas categóricas:"
        )

        for coluna in categoricas.columns:

            valores = (
                df[coluna]
                .value_counts(dropna=False)
                .head(10)
            )

            resumo.append(
                f"\nColuna: {coluna}"
            )

            resumo.append(
                valores.to_string()
            )

    return "\n".join(resumo)


def preparar_treinamento(df, alvo):
    if alvo not in df.columns:
        raise ValueError("A coluna-alvo selecionada não existe.")

    dados = df.dropna(subset=[alvo]).copy()
    if len(dados) < 20:
        raise ValueError("São necessárias pelo menos 20 linhas com valor na coluna-alvo.")

    # Evita que identificadores únicos e textos livres explodam a codificação.
    recursos = dados.drop(columns=[alvo])
    descartadas = []
    for coluna in recursos.columns:
        proporcao_unicos = recursos[coluna].nunique(dropna=True) / max(len(recursos), 1)
        if (
            pd.api.types.is_string_dtype(recursos[coluna])
            and proporcao_unicos > 0.8
        ):
            descartadas.append(coluna)
    recursos = recursos.drop(columns=descartadas)
    recursos = recursos.dropna(axis=1, how="all")
    if recursos.shape[1] == 0:
        raise ValueError("Não há colunas preditoras utilizáveis após a preparação.")

    alvo_dados = dados[alvo]
    unicos = alvo_dados.nunique(dropna=True)
    classificacao = (
        not pd.api.types.is_numeric_dtype(alvo_dados)
        or unicos <= min(30, max(2, int(len(alvo_dados) * 0.05)))
    )
    if classificacao and unicos < 2:
        raise ValueError("A coluna-alvo precisa ter pelo menos duas classes.")

    numericas = recursos.select_dtypes(include="number").columns.tolist()
    categoricas = [c for c in recursos.columns if c not in numericas]
    transformadores = []
    if numericas:
        transformadores.append(("numericas", Pipeline([
            ("imputar", SimpleImputer(strategy="median")),
            ("escalar", StandardScaler())
        ]), numericas))
    if categoricas:
        transformadores.append(("categoricas", Pipeline([
            ("imputar", SimpleImputer(strategy="most_frequent")),
            ("codificar", OneHotEncoder(handle_unknown="ignore", max_categories=50))
        ]), categoricas))

    return recursos, alvo_dados, ColumnTransformer(transformadores), classificacao, descartadas


@app.route("/")
def inicio():
    return render_template("index.html")


@app.route("/saude")
def saude():
    return jsonify({"status": "ok", "configuracao": settings.public_status()})


def diagnostico_qualidade(df):
    total_linhas, total_colunas = len(df), len(df.columns)
    total_nulos = int(df.isnull().sum().sum())
    total_celulas = total_linhas * total_colunas
    return {
        "total_colunas": total_colunas,
        "total_nulos": total_nulos,
        "percentual_nulos": round(total_nulos / total_celulas * 100, 2) if total_celulas else 0,
        "duplicados": int(df.duplicated().sum()),
        "colunas_vazias": [str(c) for c in df.columns if df[c].isnull().all()],
        "nulos_por_coluna": {str(c): int(v) for c, v in df.isnull().sum().items() if v > 0},
        "tipos": {str(c): str(t) for c, t in df.dtypes.items()}
    }


@app.route("/upload", methods=["POST"])
def upload():
    conversa_id = request.form.get("conversa_id", type=int)
    if not conversa_id or obter_conversa(conversa_id) is None:
        return jsonify({"erro": "Conversa inválida."}), 400
    ativa = obter_banco().execute("SELECT 1 FROM tarefas WHERE conversa_id=? AND status IN ('pendente','processando')", (conversa_id,)).fetchone()
    if ativa: return jsonify({"erro": "Aguarde ou cancele o treinamento antes de substituir a base."}), 409
    arquivo = request.files.get("arquivo")
    if not arquivo or not arquivo.filename:
        return jsonify({"erro": "Selecione um arquivo."}), 400
    if not extensao_permitida(arquivo.filename):
        return jsonify({"erro": "Formato inválido. Envie somente CSV ou XLSX."}), 400

    nome_original = arquivo.filename
    extensao = secure_filename(nome_original).rsplit(".", 1)[1].lower()
    caminho = os.path.join(UPLOAD_FOLDER, f"{conversa_id}_{uuid.uuid4().hex}.{extensao}")
    try:
        arquivo.save(caminho)
        df = ler_dataframe(caminho, extensao)
        if df.empty:
            raise ValueError("O arquivo está vazio.")
        if len(df) > settings.max_rows or len(df.columns) > settings.max_columns:
            raise ValueError(
                f"A base excede o limite de {settings.max_rows:,} linhas "
                f"ou {settings.max_columns:,} colunas."
            )
        qualidade = diagnostico_qualidade(df)
        banco = obter_banco()
        antigos = banco.execute("SELECT caminho FROM arquivos WHERE conversa_id = ?", (conversa_id,)).fetchall()
        modelos = banco.execute("SELECT caminho FROM modelos WHERE conversa_id = ?", (conversa_id,)).fetchall()
        banco.execute("DELETE FROM arquivos WHERE conversa_id = ?", (conversa_id,))
        cursor = banco.execute("""
            INSERT INTO arquivos (conversa_id, nome_original, caminho, extensao, tamanho,
                                  linhas, colunas, resumo_json, criada_em)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (conversa_id, nome_original, caminho, extensao, os.path.getsize(caminho),
              len(df), len(df.columns), json.dumps(qualidade, ensure_ascii=False), agora_iso()))
        banco.commit()
        for item in antigos:
            if item["caminho"] != caminho:
                remover_arquivo_seguro(item["caminho"], UPLOAD_FOLDER)
        for item in modelos:
            remover_arquivo_seguro(item["caminho"], MODEL_FOLDER)
        return jsonify({"mensagem": f"Arquivo {nome_original} carregado com sucesso!",
                        "arquivo_id": cursor.lastrowid, "linhas": len(df),
                        "colunas": [str(c) for c in df.columns], "qualidade": qualidade})
    except (ValueError, pd.errors.EmptyDataError) as erro:
        if os.path.isfile(caminho): os.remove(caminho)
        return jsonify({"erro": str(erro) or "O arquivo não possui dados válidos."}), 400
    except Exception as erro:
        if os.path.isfile(caminho): os.remove(caminho)
        logger.exception("Falha no upload", extra={"conversa_id": conversa_id})
        return jsonify({"erro": "Não foi possível processar o arquivo."}), 500


def executar_treinamento(tarefa_id, conversa_id, alvo):
    with app.app_context():
        banco = obter_banco()
        try:
            estado = banco.execute("SELECT status FROM tarefas WHERE id=?", (tarefa_id,)).fetchone()
            if estado is None or estado["status"] == "cancelada":
                return
            banco.execute("UPDATE tarefas SET status='processando', progresso=10, atualizada_em=? WHERE id=?",
                          (agora_iso(), tarefa_id)); banco.commit()
            tarefa = banco.execute("SELECT requisicao_json FROM tarefas WHERE id=?", (tarefa_id,)).fetchone()
            requisicao = json.loads(tarefa["requisicao_json"] or "{}")
            arquivo_id_esperado = requisicao.get("arquivo_id")
            arquivo = arquivo_da_conversa(conversa_id, banco)
            if arquivo is None: raise ValueError("Envie uma base antes de treinar.")
            if arquivo_id_esperado and arquivo["id"] != arquivo_id_esperado:
                raise ValueError("A base foi substituída durante o treinamento.")
            df = carregar_dataframe(arquivo)
            x, y, preparador, classificacao, descartadas = preparar_treinamento(df, alvo)
            contagens = y.value_counts()
            avisos = []
            alvo_normalizado = str(alvo).strip().lower().replace(" ", "_")
            suspeitas_vazamento = [
                str(coluna) for coluna in x.columns
                if alvo_normalizado in str(coluna).strip().lower().replace(" ", "_")
            ]
            if pd.api.types.is_numeric_dtype(y):
                for coluna in x.select_dtypes(include="number").columns:
                    correlacao = x[coluna].corr(y)
                    if pd.notna(correlacao) and abs(correlacao) > .98:
                        suspeitas_vazamento.append(str(coluna))
            suspeitas_vazamento = sorted(set(suspeitas_vazamento))
            if len(y) < 100: avisos.append("Base pequena: interprete as métricas com cautela.")
            if classificacao and contagens.min() / contagens.max() < .25:
                avisos.append("Classes desbalanceadas detectadas.")
            if suspeitas_vazamento:
                avisos.append("Possível vazamento de dados detectado; revise as colunas sinalizadas.")
            estratificar = y if classificacao and contagens.min() >= 2 else None
            x_tr, x_te, y_tr, y_te = train_test_split(x, y, test_size=.2, random_state=42, stratify=estratificar)
            modelos = ({
                "Regressão logística": LogisticRegression(max_iter=1500, class_weight="balanced"),
                "Floresta aleatória": RandomForestClassifier(n_estimators=200, random_state=42, n_jobs=-1, class_weight="balanced")
            } if classificacao else {
                "Regressão linear": LinearRegression(),
                "Floresta aleatória": RandomForestRegressor(n_estimators=200, random_state=42, n_jobs=-1)
            })
            resultados, melhor, melhor_nome, melhor_score = [], None, None, float("-inf")
            scoring = "f1_weighted" if classificacao else "r2"
            folds = min(5, int(contagens.min()) if classificacao else len(y) // 5)
            folds = max(2, folds)
            cv = StratifiedKFold(folds, shuffle=True, random_state=42) if classificacao else KFold(folds, shuffle=True, random_state=42)
            baseline = DummyClassifier(strategy="most_frequent") if classificacao else DummyRegressor(strategy="mean")
            baseline_pipeline = Pipeline([("preparacao", preparador), ("modelo", baseline)])
            baseline_score = float(cross_val_score(baseline_pipeline, x, y, cv=cv, scoring=scoring).mean())
            for indice, (nome, estimador) in enumerate(modelos.items()):
                status = banco.execute("SELECT status FROM tarefas WHERE id=?", (tarefa_id,)).fetchone()
                if status and status["status"] == "cancelada":
                    return
                pipeline = Pipeline([("preparacao", preparador), ("modelo", estimador)])
                pipeline.fit(x_tr, y_tr); previsto = pipeline.predict(x_te)
                cv_scores = cross_val_score(pipeline, x, y, cv=cv, scoring=scoring, n_jobs=1)
                if classificacao:
                    metricas = {"acuracia": accuracy_score(y_te, previsto),
                                "f1": f1_score(y_te, previsto, average="weighted", zero_division=0),
                                "precisao": precision_score(y_te, previsto, average="weighted", zero_division=0),
                                "recall": recall_score(y_te, previsto, average="weighted", zero_division=0),
                                "matriz_confusao": confusion_matrix(y_te, previsto).tolist(),
                                "validacao_cruzada": float(cv_scores.mean())}
                else:
                    metricas = {"r2": r2_score(y_te, previsto), "mae": mean_absolute_error(y_te, previsto),
                                "rmse": mean_squared_error(y_te, previsto) ** .5,
                                "validacao_cruzada": float(cv_scores.mean())}
                score = float(cv_scores.mean())
                metricas = {k: (round(float(v), 4) if not isinstance(v, list) else v) for k, v in metricas.items()}
                resultados.append({"modelo": nome, "metricas": metricas})
                if score > melhor_score: melhor, melhor_nome, melhor_score = pipeline, nome, score
                banco.execute("UPDATE tarefas SET progresso=?, atualizada_em=? WHERE id=?",
                              (45 + indice * 35, agora_iso(), tarefa_id)); banco.commit()
            resultados.sort(key=lambda r: r["metricas"]["validacao_cruzada"], reverse=True)
            importancia = []
            estimador_final = melhor.named_steps["modelo"]
            if hasattr(estimador_final, "feature_importances_"):
                nomes = melhor.named_steps["preparacao"].get_feature_names_out()
                importancia = sorted(
                    ({"recurso": str(nome), "importancia": round(float(valor), 6)}
                     for nome, valor in zip(nomes, estimador_final.feature_importances_)),
                    key=lambda item: item["importancia"], reverse=True
                )[:15]
            caminho_modelo = os.path.join(MODEL_FOLDER, f"{conversa_id}_{uuid.uuid4().hex}.joblib")
            pacote = {"pipeline": melhor, "colunas": list(x.columns), "tipos": {str(c): str(x[c].dtype) for c in x.columns}, "alvo": alvo,
                      "tipo": "classificação" if classificacao else "regressão",
                      "versoes": {"python": platform.python_version(), "sklearn": sklearn.__version__},
                      "treinado_em": agora_iso(), "dataset_sha256": sha256_arquivo(arquivo["caminho"])}
            joblib.dump(pacote, caminho_modelo)
            hash_modelo = sha256_arquivo(caminho_modelo)
            resultado = {"tipo": pacote["tipo"], "alvo": alvo, "linhas_treino": len(x_tr),
                         "linhas_teste": len(x_te), "recursos": x.shape[1], "avisos": avisos,
                         "colunas_descartadas": descartadas, "melhor_modelo": melhor_nome,
                         "baseline_validacao_cruzada": round(baseline_score, 4),
                         "possivel_vazamento": suspeitas_vazamento,
                         "importancia_recursos": importancia, "versoes": pacote["versoes"],
                         "resultados": resultados}
            atual = arquivo_da_conversa(conversa_id, banco)
            if atual is None or atual["id"] != arquivo["id"]:
                remover_arquivo_seguro(caminho_modelo, MODEL_FOLDER)
                raise ValueError("A base mudou antes da conclusão; o modelo foi descartado.")
            antigos = banco.execute("SELECT caminho FROM modelos WHERE conversa_id=?", (conversa_id,)).fetchall()
            banco.execute("DELETE FROM modelos WHERE conversa_id=?", (conversa_id,))
            banco.execute("""INSERT INTO modelos (conversa_id, arquivo_id, alvo, tipo, nome, caminho,
                              recursos_json, metricas_json, criada_em, sha256) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                          (conversa_id, arquivo["id"], alvo, pacote["tipo"], melhor_nome, caminho_modelo,
                           json.dumps(list(x.columns), ensure_ascii=False), json.dumps(resultado, ensure_ascii=False), agora_iso(), hash_modelo))
            banco.execute("UPDATE tarefas SET status='concluida', progresso=100, resultado_json=?, atualizada_em=? WHERE id=?",
                          (json.dumps(resultado, ensure_ascii=False), agora_iso(), tarefa_id)); banco.commit()
            for item in antigos: remover_arquivo_seguro(item["caminho"], MODEL_FOLDER)
        except Exception as erro:
            logger.exception("Falha no treinamento", extra={"tarefa_id": tarefa_id, "conversa_id": conversa_id})
            banco.execute("UPDATE tarefas SET status='falhou', erro=?, atualizada_em=? WHERE id=?",
                          (str(erro), agora_iso(), tarefa_id)); banco.commit()
        finally:
            task_slots.release()


@app.route("/treinar", methods=["POST"])
def treinar():
    dados = request.get_json(silent=True) or {}; conversa_id = dados.get("conversa_id"); alvo = str(dados.get("alvo", "")).strip()
    if not isinstance(conversa_id, int) or obter_conversa(conversa_id) is None:
        return jsonify({"erro": "Conversa inválida."}), 400
    if not alvo: return jsonify({"erro": "Selecione a coluna que deseja prever."}), 400
    if arquivo_da_conversa(conversa_id) is None: return jsonify({"erro": "Envie uma base antes de treinar."}), 400
    if not task_slots.acquire(blocking=False):
        return jsonify({"erro": "A fila de treinamento está cheia. Tente novamente mais tarde."}), 503
    tarefa_id = uuid.uuid4().hex; agora = agora_iso(); banco = obter_banco()
    arquivo_id = arquivo_da_conversa(conversa_id)["id"]
    try:
        banco.execute("INSERT INTO tarefas (id, conversa_id, tipo, status, progresso, criada_em, atualizada_em, requisicao_json) VALUES (?,?,'treinamento','pendente',0,?,?,?)",
                      (tarefa_id, conversa_id, agora, agora, json.dumps({"alvo": alvo, "arquivo_id": arquivo_id}, ensure_ascii=False)))
        banco.commit()
        executor.submit(executar_treinamento, tarefa_id, conversa_id, alvo)
    except Exception:
        task_slots.release()
        raise
    return jsonify({"tarefa_id": tarefa_id, "status": "pendente"}), 202


@app.route("/tarefas/<tarefa_id>")
def consultar_tarefa(tarefa_id):
    tarefa = obter_banco().execute("SELECT * FROM tarefas WHERE id=?", (tarefa_id,)).fetchone()
    if tarefa is None: return jsonify({"erro": "Tarefa não encontrada."}), 404
    dados = dict(tarefa); dados["resultado"] = json.loads(dados.pop("resultado_json")) if dados["resultado_json"] else None
    return jsonify(dados)


@app.route("/tarefas/<tarefa_id>", methods=["DELETE"])
def cancelar_tarefa(tarefa_id):
    banco = obter_banco()
    cursor = banco.execute(
        "UPDATE tarefas SET status='cancelada', atualizada_em=? WHERE id=? AND status IN ('pendente','processando')",
        (agora_iso(), tarefa_id),
    )
    banco.commit()
    if cursor.rowcount == 0:
        return jsonify({"erro": "Tarefa não encontrada ou já finalizada."}), 409
    return "", 204


@app.route("/prever", methods=["POST"])
def prever():
    conversa_id = request.form.get("conversa_id", type=int); arquivo = request.files.get("arquivo")
    if not conversa_id or obter_conversa(conversa_id) is None: return jsonify({"erro": "Conversa inválida."}), 400
    modelo = modelo_da_conversa(conversa_id)
    if modelo is None or not os.path.isfile(modelo["caminho"]): return jsonify({"erro": "Treine um modelo antes de prever."}), 400
    if not arquivo or not extensao_permitida(arquivo.filename): return jsonify({"erro": "Envie um CSV ou XLSX válido."}), 400
    try:
        extensao = "csv" if arquivo.filename.lower().endswith(".csv") else "xlsx"
        entrada = ler_dataframe(arquivo, extensao)
        if not modelo["sha256"] or not hmac.compare_digest(modelo["sha256"], sha256_arquivo(modelo["caminho"])):
            return jsonify({"erro": "O artefato do modelo falhou na verificação de integridade."}), 409
        pacote = joblib.load(modelo["caminho"]); ausentes = [c for c in pacote["colunas"] if c not in entrada.columns]
        if ausentes: return jsonify({"erro": "Faltam colunas: " + ", ".join(map(str, ausentes))}), 400
        incompatíveis = []
        for coluna in pacote["colunas"]:
            esperado_numerico = pacote.get("tipos", {}).get(str(coluna), "").startswith(("int", "float"))
            if esperado_numerico and not pd.api.types.is_numeric_dtype(entrada[coluna]):
                incompatíveis.append(str(coluna))
        if incompatíveis:
            return jsonify({"erro": "Tipos incompatíveis nas colunas: " + ", ".join(incompatíveis)}), 400
        previsoes = pacote["pipeline"].predict(entrada[pacote["colunas"]]); saida = entrada.copy()
        saida[f"previsao_{pacote['alvo']}"] = previsoes
        if pacote["tipo"] == "classificação" and hasattr(pacote["pipeline"], "predict_proba"):
            probabilidades = pacote["pipeline"].predict_proba(entrada[pacote["colunas"]])
            saida["confianca_previsao"] = probabilidades.max(axis=1)
        token = uuid.uuid4().hex; caminho = os.path.join(PREDICTION_FOLDER, f"{token}.csv")
        neutralizar_formulas_csv(saida).to_csv(caminho, index=False, encoding="utf-8-sig")
        agora = datetime.now(timezone.utc)
        return jsonify({"mensagem": f"{len(saida)} previsões geradas.", "download": f"/previsoes/{token}"})
    except Exception as erro:
        logger.exception("Falha na previsão", extra={"conversa_id": conversa_id})
        return jsonify({"erro": "Não foi possível gerar as previsões."}), 500


@app.route("/previsoes/<token>")
def baixar_previsao(token):
    limpar_previsoes_expiradas()
    if not token.isalnum(): return jsonify({"erro": "Arquivo inválido."}), 400
    caminho = os.path.join(PREDICTION_FOLDER, f"{token}.csv")
    if not os.path.isfile(caminho): return jsonify({"erro": "Previsão não encontrada."}), 404
    return send_file(caminho, as_attachment=True, download_name="previsoes.csv")


@app.route("/perguntar", methods=["POST"])
def perguntar():
    dados = request.get_json(silent=True) or {}; mensagem = str(dados.get("mensagem", "")).strip(); conversa_id = dados.get("conversa_id")
    if not isinstance(conversa_id, int) or obter_conversa(conversa_id) is None: return jsonify({"erro": "Conversa inválida."}), 400
    if not mensagem: return jsonify({"erro": "Digite uma mensagem."}), 400
    if len(mensagem) > settings.max_message_chars: return jsonify({"erro": "Sua mensagem é muito grande."}), 400
    mensagem_id = salvar_mensagem(conversa_id, "usuario", mensagem, "processando")
    try:
        registro = arquivo_da_conversa(conversa_id); contexto_dados = ""
        if registro:
            df = carregar_dataframe(registro)
            df_prompt, sensiveis = dataframe_privado_para_prompt(df)
            contexto_dados = f"\nBase carregada.\n{gerar_resumo_dataframe(df_prompt)}\n"
            if sensiveis:
                contexto_dados += "\nAlguns nomes de colunas foram ocultados por privacidade.\n"
        historico = obter_banco().execute("SELECT autor, conteudo FROM mensagens WHERE conversa_id=? ORDER BY id DESC LIMIT 12", (conversa_id,)).fetchall()
        contexto = "Você é um assistente de análise de dados. Responda em português, sem inventar valores.\n" + contexto_dados
        contexto += "\nHistórico:\n" + "\n".join(f"{m['autor']}: {m['conteudo']}" for m in reversed(historico))
        resposta_texto = ai.send_message(settings, contexto)
        salvar_mensagem(conversa_id, "agente", resposta_texto)
        banco = obter_banco(); banco.execute("UPDATE mensagens SET status='concluida', erro=NULL WHERE id=?", (mensagem_id,))
        total = banco.execute("SELECT COUNT(*) FROM mensagens WHERE conversa_id=?", (conversa_id,)).fetchone()[0]; titulo = None
        if total == 2:
            titulo = mensagem[:55] + ("…" if len(mensagem) > 55 else ""); banco.execute("UPDATE conversas SET titulo=? WHERE id=?", (titulo, conversa_id))
        banco.commit(); return jsonify({"resposta": resposta_texto, "titulo": titulo})
    except Exception as erro:
        texto = str(erro); codigo = 500; mensagem_erro = "Não foi possível obter resposta da IA."
        if "429" in texto or "RESOURCE_EXHAUSTED" in texto: codigo, mensagem_erro = 429, "Limite da IA atingido. Aguarde e tente novamente."
        elif "401" in texto or "API_KEY" in texto: codigo, mensagem_erro = 503, "A chave da IA está inválida ou ausente."
        elif "timeout" in texto.lower(): codigo, mensagem_erro = 504, "A IA demorou demais para responder. Tente novamente."
        banco = obter_banco(); banco.execute("UPDATE mensagens SET status='falhou', erro=? WHERE id=?", (mensagem_erro, mensagem_id)); banco.commit()
        logger.warning("Falha na IA conversa_id=%s tipo=%s", conversa_id, type(erro).__name__)
        return jsonify({"erro": mensagem_erro, "mensagem_id": mensagem_id}), codigo


@app.route("/mensagens/<int:mensagem_id>/tentar-novamente", methods=["POST"])
def tentar_novamente(mensagem_id):
    item = obter_banco().execute("SELECT conversa_id, conteudo, status FROM mensagens WHERE id=? AND autor='usuario'", (mensagem_id,)).fetchone()
    if item is None: return jsonify({"erro": "Mensagem não encontrada."}), 404
    obter_banco().execute("DELETE FROM mensagens WHERE id=?", (mensagem_id,)); obter_banco().commit()
    with app.test_request_context(json={"conversa_id": item["conversa_id"], "mensagem": item["conteudo"]}):
        return perguntar()


@app.errorhandler(413)
def arquivo_muito_grande(erro):

    return jsonify({
        "erro":
            "Arquivo muito grande. O limite é 10 MB."
    }), 413


iniciar_banco()


def recuperar_tarefas_pendentes():
    with app.app_context():
        banco = obter_banco()
        tarefas = banco.execute(
            "SELECT id, conversa_id, requisicao_json FROM tarefas WHERE status IN ('pendente','processando')"
        ).fetchall()
        for tarefa in tarefas:
            requisicao = json.loads(tarefa["requisicao_json"] or "{}")
            alvo = requisicao.get("alvo")
            if alvo:
                if not task_slots.acquire(blocking=False):
                    banco.execute("UPDATE tarefas SET status='falhou', erro=?, atualizada_em=? WHERE id=?",
                                  ("Fila cheia durante a recuperação.", agora_iso(), tarefa["id"]))
                    continue
                banco.execute(
                    "UPDATE tarefas SET status='pendente', progresso=0, atualizada_em=? WHERE id=?",
                    (agora_iso(), tarefa["id"]),
                )
                executor.submit(executar_treinamento, tarefa["id"], tarefa["conversa_id"], alvo)
            else:
                banco.execute(
                    "UPDATE tarefas SET status='falhou', erro=?, atualizada_em=? WHERE id=?",
                    ("Tarefa antiga sem parâmetros para retomada.", agora_iso(), tarefa["id"]),
                )
        banco.commit()


recuperar_tarefas_pendentes()


if __name__ == "__main__":
    host = os.getenv("ATLAS_HOST", "127.0.0.1")
    debug = os.getenv("FLASK_DEBUG", "false").lower() == "true"
    if host not in {"127.0.0.1", "localhost", "::1"} and debug:
        raise RuntimeError("Uso em rede exige FLASK_DEBUG=false.")
    app.run(host=host, debug=debug)
